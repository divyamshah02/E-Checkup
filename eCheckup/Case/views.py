from rest_framework import viewsets, status
from rest_framework.response import Response
from django.utils import timezone
from django.db.models import Q
from django.http import JsonResponse

from .models import *
from .serializers import *

from UserDetail.models import User
from UserDetail.serializers import UserSerializer
from LIC.models import BranchOffice, DivisionalOffice, RegionalOffice, HeadOffice
from utils.decorators import check_authentication, handle_exceptions
from utils.Notification_System import send_welcome, send_scheduled, send_medical_email, send_feedback

import calendar
from datetime import datetime, timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.timezone import make_aware
from utils.handle_s3_bucket import upload_file_to_s3
from django.utils.timezone import localtime

import random, string
import time as time_fun

class CaseViewSet(viewsets.ViewSet):

    @check_authentication(required_role=['admin', 'hod', 'coordinator'])
    @handle_exceptions
    def create(self, request):
        data = request.data
        case_type = data.get("case_type")

        if not case_type:
            return Response({"error": "Missing case_type."}, status=400)

        prefix = {
            'vmer': 'VM',
            'dc_visit': 'DC',
            'online': 'ON',
            'both': 'BT'
        }.get(case_type, 'XX')

        while True:
            generated_id = prefix + ''.join(random.choices(string.digits, k=10))
            if not Case.objects.filter(case_id=generated_id).exists():
                break

        data['case_id'] = generated_id
        data['created_by'] = request.user.user_id
        data['status'] = 'created'

        serializer = CaseSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            CaseActionLog.objects.create(case_id=generated_id, action_by=request.user.user_id, action="Case Created")
            send_welcome(phone=serializer.data.get('holder_phone'), recipient_email=serializer.data.get('holder_email'))
            return Response({"success": True, "data": serializer.data}, status=201)
        print(serializer.errors)
        return Response({"success": False, "error": serializer.errors}, status=400)

    @check_authentication()
    @handle_exceptions
    def extra_divyam(self, request):
        user = request.user
        user_role = user.role

        case_id = request.query_params.get('case_id')
        case_type = request.query_params.get('type')
        is_dashboard = request.query_params.get('is_dashboard', False)

        # Single Case View
        if case_id:
            case = Case.objects.filter(case_id=case_id, is_active=True).first()
            if not case:
                return Response({"error": "Case not found."}, status=404)
            serializer = CaseDetailSerializer(case)
            return Response({"success": True, "data": serializer.data})

        # Admin view by case_type
        if case_type and user_role == 'admin':
            cases = Case.objects.filter(case_type=case_type, is_active=True)
            serializer = CaseSerializer(cases, many=True)
            return Response({"success": True, "data": serializer.data})

        all_cases = Case.objects.none()
        pending_cases = Case.objects.none()
        completed_cases = Case.objects.none()
        ninety_days_ago = timezone.now() - timedelta(days=90)
        if user_role == 'hod':
            if is_dashboard == 'dashboard':
                all_cases = Case.objects.filter(is_active=True, created_at__gte=ninety_days_ago)
            else:
                all_cases = Case.objects.filter(is_active=True)
            pending_cases = all_cases.exclude(status__in=['completed'])
            completed_cases = all_cases.filter(status='completed')

        elif user_role == 'coordinator':
            if is_dashboard == 'dashboard':
                all_cases = Case.objects.filter(assigned_coordinator_id=user.user_id, is_active=True, created_at__gte=ninety_days_ago)
            else:
                all_cases = Case.objects.filter(assigned_coordinator_id=user.user_id, is_active=True)
            pending_cases = all_cases.exclude(status='submitted_to_lic')
            completed_cases = all_cases.filter(status='submitted_to_lic')

        elif user_role == 'telecaller':
            if is_dashboard == 'dashboard':
                all_cases = Case.objects.filter(assigned_telecaller_id=user.user_id, is_active=True, created_at__gte=ninety_days_ago)
            else:
                all_cases = Case.objects.filter(assigned_telecaller_id=user.user_id, is_active=True)
            pending_cases = all_cases.filter(status='assigned')
            completed_cases = all_cases.exclude(status='assigned')

        elif user_role == 'vmer_med_co':
            if is_dashboard == 'dashboard':
                all_cases = Case.objects.filter(assigned_vmer_med_co_id=user.user_id, is_active=True, created_at__gte=ninety_days_ago)
            else:
                all_cases = Case.objects.filter(assigned_vmer_med_co_id=user.user_id, is_active=True)
            pending_cases = all_cases.filter(status='scheduled')
            completed_cases = all_cases.exclude(status='scheduled')

        elif user_role == 'diagnostic_center':
            if is_dashboard == 'dashboard':
                all_cases = Case.objects.filter(assigned_dc_id=user.user_id, is_active=True, created_at__gte=ninety_days_ago)
            else:
                all_cases = Case.objects.filter(assigned_dc_id=user.user_id, is_active=True)
            pending_cases = all_cases.filter(status='scheduled')
            completed_cases = all_cases.exclude(status='scheduled')

        elif user_role == 'lic':
            if is_dashboard == 'dashboard':
                all_cases = Case.objects.filter(is_active=True, created_at__gte=ninety_days_ago)
            else:
                all_cases = Case.objects.filter(is_active=True)
            # Filter by cases where this LIC user is in that hierarchy — omitted logic, assumed handled elsewhere
            pending_cases = all_cases.filter(status='submitted_to_lic')
            completed_cases = all_cases.exclude(status='submitted_to_lic')

        else:
            return Response({"error": "Unauthorized role"}, status=403)

        return Response({
            "success": True,
            "data": {
                "all_cases": CaseSerializer(all_cases, many=True).data[::-1],
                "pending_cases": CaseSerializer(pending_cases, many=True).data,
                "completed_cases": CaseSerializer(completed_cases, many=True).data,
            }
        })


    @check_authentication()
    @handle_exceptions
    def list(self, request):
        user = request.user
        user_role = user.role

        case_id = request.query_params.get('case_id')
        case_type = request.query_params.get('type')
        is_dashboard = request.query_params.get('is_dashboard', False)

        # Single Case View (unchanged behaviour)
        if case_id:
            case = Case.objects.filter(case_id=case_id, is_active=True).first()
            if not case:
                return Response({"error": "Case not found."}, status=404)
            serializer = CaseDetailSerializer(case)
            return Response({"success": True, "data": serializer.data})

        # Admin view by case_type (unchanged)
        if case_type and user_role == 'admin':
            cases = Case.objects.filter(case_type=case_type, is_active=True)
        else:
            ninety_days_ago = timezone.now() - timedelta(days=90)
            base_filters = Q(is_active=True)
            if is_dashboard == 'dashboard':
                base_filters &= Q(created_at__gte=ninety_days_ago)

            # Build queryset based on role (same filters you used)
            if user_role == 'hod':
                cases = Case.objects.filter(base_filters)
                pending_statuses = None  # means all except 'completed'
                completed_statuses = {'completed'}

            elif user_role == 'coordinator':
                cases = Case.objects.filter(base_filters & Q(assigned_coordinator_id=user.user_id))
                pending_statuses = None  # all except 'submitted_to_lic'
                completed_statuses = {'submitted_to_lic'}

            elif user_role == 'telecaller':
                cases = Case.objects.filter(base_filters & Q(assigned_telecaller_id=user.user_id))
                pending_statuses = {'assigned'}
                completed_statuses = None  # all except 'assigned'

            elif user_role == 'vmer_med_co':
                cases = Case.objects.filter(base_filters & Q(assigned_vmer_med_co_id=user.user_id))
                pending_statuses = {'scheduled'}
                completed_statuses = None  # all except 'scheduled'

            elif user_role == 'diagnostic_center':
                cases = Case.objects.filter(base_filters & Q(assigned_dc_id=user.user_id))
                pending_statuses = {'scheduled'}
                completed_statuses = None  # all except 'scheduled'

            elif user_role == 'lic':
                cases = Case.objects.filter(base_filters)
                pending_statuses = {'submitted_to_lic'}
                completed_statuses = None  # all except 'submitted_to_lic'

            else:
                return Response({"error": "Unauthorized role"}, status=403)

        # --- BULK user fetch (one query) ---
        # Evaluate queryset only once while collecting ids. This will not create extra DB hits per case.
        user_ids = set()
        for c in cases:
            # keep identical fields you already rely upon
            user_ids.add(c.assigned_coordinator_id)
            if c.assigned_telecaller_id:
                user_ids.add(c.assigned_telecaller_id)
            if c.assigned_dc_id:
                user_ids.add(c.assigned_dc_id)
            if c.assigned_vmer_med_co_id:
                user_ids.add(c.assigned_vmer_med_co_id)
        user_ids.discard(None)

        # Single DB hit to fetch all users referenced
        user_map = {}
        if user_ids:
            # only fetch user_id and name to reduce load
            qs_users = User.objects.filter(user_id__in=user_ids).only('user_id', 'name')
            user_map = {u.user_id: u for u in qs_users}  # store object to preserve .name access like before

        # Serialize once, add names from the user_map (no DB hits here)
        serialized = CaseSerializer(cases, many=True, context={'user_map': user_map}).data

        # Build pending/completed lists using the status logic assigned above (replicates your previous behaviour)
        def is_pending(case_obj):
            st = case_obj.get('status')
            if pending_statuses is None:
                # pending = all except completed OR except specific completed_status depending on role.
                if user_role == 'hod':
                    return st != 'completed'
                if user_role == 'coordinator':
                    return st != 'submitted_to_lic'
                return True  # fallback
            else:
                return st in pending_statuses

        def is_completed(case_obj):
            st = case_obj.get('status')
            if completed_statuses is None:
                # completed = all except the pending status set for certain roles
                if user_role == 'telecaller':
                    return st != 'assigned'
                if user_role in ('vmer_med_co', 'diagnostic_center'):
                    return st != 'scheduled'
                if user_role == 'lic':
                    return st != 'submitted_to_lic'
                return False
            else:
                return st in completed_statuses

        pending_cases_data = [c for c in serialized if is_pending(c)]
        completed_cases_data = [c for c in serialized if is_completed(c)]

        # Return (keeps same structure and order, reversed all_cases like before)
        return Response({
            "success": True,
            "data": {
                "all_cases": serialized[::-1],
                "pending_cases": pending_cases_data,
                "completed_cases": completed_cases_data,
            }
        })

    @check_authentication(required_role=['admin', 'hod', 'coordinator', 'telecaller', 'vmer_med_co', 'diagnostic_center'])
    @handle_exceptions
    def update(self, request, pk=None):
        case_id = request.data.get("case_id")
        case = Case.objects.filter(case_id=case_id, is_active=True).first()
        if not case:
            return Response({"error": "Invalid case_id."}, status=404)

        serializer = CaseSerializer(case, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            if request.data.get('status') == 'submitted_to_lic':
                send_feedback(name=case.holder_name, feedback_form_link='www.google.com', recipient_email=case.holder_email, phone=case.holder_phone)
                CaseActionLog.objects.create(case_id=case.case_id, action_by=request.user.user_id, action="Case Submitted to LIC")
            elif request.data.get('status') == 'issue':
                CaseActionLog.objects.create(case_id=case.case_id, action_by=request.user.user_id, action="Issue Raised")
            else:
                CaseActionLog.objects.create(case_id=case.case_id, action_by=request.user.user_id, action="Case Updated")

            return Response({"success": True, "data": serializer.data})
        return Response({"error": serializer.errors}, status=400)

    @check_authentication(required_role=['admin', 'hod'])
    @handle_exceptions
    def destroy(self, request, pk=None):
        case_id = request.data.get("case_id")
        case = Case.objects.filter(case_id=case_id).first()
        if not case:
            return Response({"error": "Case not found."}, status=404)
        case.is_active = False
        case.save()
        CaseActionLog.objects.create(case_id=case.case_id, action_by=request.user.user_id, action="Case Deleted")
        return Response({"success": True, "message": "Case soft-deleted."})


class CreateCaseFromExcelViewSet(viewsets.ViewSet):

    @check_authentication(required_role=['admin', 'hod', 'coordinator'])
    @handle_exceptions
    def create(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file uploaded."}, status=400)

        try:
            import pandas as pd
            df = pd.read_excel(excel_file)
        except Exception as e:
            return Response({"error": f"Failed to read Excel file: {str(e)}"}, status=400)

        required_columns = [
            "case_type", "policy_type", "policy_number", "insurance_company", "priority", "due_date",
            "holder_name", "holder_phone", "holder_email", "ins_office_code",
            "lic_agent", "assigned_coordinator_email", "payment_method", "lic_gst_no",
            "lic_type", "intimation_date", "holder_dob", "holder_gender",
            "holder_address", "holder_state", "holder_city", "holder_pincode",
            "proposed_sum_insured", "sum_insured_under_consideration", "tests",
            "special_instructions"
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=400)

        created_count = 0
        failed_cases = []

        # Define allowed choices
        valid_case_types = ['vmer', 'dc_visit', 'online', 'both']
        valid_policy_types = ['new', 'revival']
        valid_priorities = ['normal', 'urgent']
        valid_payment_methods = ['lic', 'self']
        valid_lic_types = ['urban', 'rural']
        valid_genders = ['M', 'F']
        valid_insurance_company = ['LIC', 'TATA']

        for col in df.select_dtypes(include=['datetime64[ns]']).columns:
            df[col] = df[col].dt.date
        for idx, row in df.iterrows():
            case_data = row.to_dict()

            # Validate case_type
            case_type = case_data.get('case_type')
            if case_type not in valid_case_types:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Invalid case_type '{case_type}'"
                })
                continue

            # Validate policy_type
            policy_type = case_data.get('policy_type')
            if policy_type not in valid_policy_types:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Invalid policy_type '{policy_type}'"
                })
                continue

            # Validate priority
            priority = case_data.get('priority')
            if priority not in valid_priorities:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Invalid priority '{priority}'"
                })
                continue

            # Validate payment_method
            payment_method = case_data.get('payment_method')
            if payment_method and payment_method not in valid_payment_methods:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Invalid payment_method '{payment_method}'"
                })
                continue

            # Validate lic_type
            lic_type = case_data.get('lic_type')
            if lic_type not in valid_lic_types:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Invalid lic_type '{lic_type}'"
                })
                continue

            # Validate holder_gender
            holder_gender = case_data.get('holder_gender')
            if holder_gender and holder_gender not in valid_genders:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Invalid holder_gender '{holder_gender}'"
                })
                continue

            # Validate holder_gender
            insurance_company = case_data.get('insurance_company')
            if insurance_company and insurance_company not in valid_insurance_company:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Invalid insurance_company '{insurance_company}'"
                })
                continue

            # Coordinator validation
            assigned_coordinator_email = case_data.get('assigned_coordinator_email')
            coordinator = User.objects.filter(
                email=assigned_coordinator_email,
                role='coordinator',
                is_active=True
            ).first()
            if not coordinator:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": f"Coordinator with email '{assigned_coordinator_email}' not found"
                })
                continue
            case_data['assigned_coordinator_id'] = coordinator.user_id

            # Handle tests
            tests_data = case_data.get('tests')
            tests = [str(test).strip() for test in str(tests_data).split('|')]
            selectedTestPrices = {}
            invalid_test = False
            for test in tests:
                testDetails = TestDetail.objects.filter(test_name=test).first()
                if not testDetails:
                    failed_cases.append({
                        "row_index": idx + 2,
                        "holder_name": case_data.get("holder_name"),
                        "reason": f"Test '{test}' not found in system"
                    })
                    invalid_test = True
                    break
                selectedTestPrices[testDetails.test_name] = {
                    'test_name': testDetails.test_name,
                    'dc_charge': testDetails.dc_charge,
                    'lic_rural_charge': testDetails.lic_rural_charge,
                    'lic_urban_charge': testDetails.lic_urban_charge,
                }
            if invalid_test:
                continue

            # Generate case_id
            prefix = {
                'vmer': 'VM',
                'dc_visit': 'DC',
                'online': 'ON',
                'both': 'BT'
            }.get(case_type, 'XX')

            while True:
                generated_id = prefix + ''.join(random.choices(string.digits, k=10))
                if not Case.objects.filter(case_id=generated_id).exists():
                    break

            case_data['tests'] = tests
            case_data['test_price'] = selectedTestPrices
            case_data['case_id'] = generated_id
            case_data['created_by'] = request.user.user_id
            case_data['status'] = 'created'

            serializer = CaseSerializer(data=case_data)
            if serializer.is_valid():
                serializer.save()
                CaseActionLog.objects.create(
                    case_id=generated_id,
                    action_by=request.user.user_id,
                    action="Case Created"
                )
                created_count += 1
                send_welcome(
                    phone=serializer.data.get('holder_phone'),
                    recipient_email=serializer.data.get('holder_email')
                )
            else:
                failed_cases.append({
                    "row_index": idx + 2,
                    "holder_name": case_data.get("holder_name"),
                    "reason": serializer.errors
                })

        report = {
            "success": True,
            "total_cases_created": created_count,
            "total_failed_cases": len(failed_cases),
            "failed_cases": failed_cases
        }

        return Response(report, status=201)


class CaseAssignmentViewSet(viewsets.ViewSet):

    @check_authentication()
    @handle_exceptions
    def create(self, request):
        case_id = request.data.get("case_id")
        assign_to = request.data.get("assign_to")
        role = request.data.get("role")
        print(assign_to)

        case = Case.objects.get(case_id=case_id, is_active=True)
        if not case:
            return Response({"error": "Invalid case_id."}, status=404)
        assign_to_obj = User.objects.filter(user_id=assign_to).first()
        if role == 'telecaller':
            case.assigned_telecaller_id = assign_to
            case.status = 'assigned'
            action = f"Assigned to Telecaller - {assign_to_obj.name}"
        elif role == 'vmer_med_co':
            case.assigned_vmer_med_co_id = assign_to
            action = f"Assigned to VMER Med Co - {assign_to_obj.name}"
        elif role == 'diagnostic_center':
            dc_details = DiagnosticCenter.objects.filter(user_id=assign_to).first()
            print(dc_details.name)
            try:
                schedule_data = Schedule.objects.filter(case_id=case_id, is_active=True).first()
                local_time = localtime(schedule_data.schedule_time)

                date = local_time.strftime("%d-%b-%Y")
                time = local_time.strftime("%I:%M %p")
                send_scheduled(
                    date=date,
                    time=time,
                    dc_name=dc_details.name,
                    address=dc_details.address,
                    gmap_link=f"https://www.google.com/maps/search/{str(dc_details.name).replace(" ", "%20")}+{str(dc_details.address).replace(" ", "%20")}+{str(dc_details.city).replace(" ", "%20")}+{str(dc_details.state).replace(" ", "%20")}",
                    contact_number="7718861051",
                    email_id="lic.pims@ericsontpa.com",
                    recipient_email=case.holder_email,
                    phone=case.holder_phone,
                )

                time_fun.sleep(5)
                send_medical_email(                    
                    recipient_email=assign_to_obj.email,
                    appointment_date=date,
                    subject=f"Appointment on {date} {case.holder_name}",
                    insurance_company="LIC OF INDIA",
                    intimation_number=case_id,
                    branch_code=case.ins_office_code,
                    proposal_number=case.policy_number,
                    client_name=case.holder_name,
                    dob=f"{case.holder_dob.day}/{case.holder_dob.month}/{case.holder_dob.year}",
                    gender=str(case.holder_gender).upper(),
                    contact_number=case.holder_phone,
                    sum_assured=case.sum_insured_under_consideration,
                    medical_test="#".join(case.tests),
                    intimation_date=f"{case.intimation_date.day}/{case.intimation_date.month}/{case.intimation_date.year}",
                    appointment_time=time,
                    visit_type="Centre visit",
                    client_address=case.holder_address
                )

            except Exception as e:
                print(e)
            case.assigned_dc_id = assign_to
            action = f"Assigned to Diagnostic Center - {dc_details.name}"
        else:
            return Response({"error": "Invalid role."}, status=400)

        case.save()
        CaseActionLog.objects.create(case_id=case.case_id, action_by=request.user.user_id, action=action)
        return Response({"success": True, "message": action})


class ScheduleViewSet(viewsets.ViewSet):

    @check_authentication(required_role=['telecaller'])
    @handle_exceptions
    def create(self, request):
        case_id = request.data.get("case_id")
        schedule_time = request.data.get("schedule_time")
        reason = request.data.get("reason", None)

        if not case_id or not schedule_time:
            return Response({"error": "Missing required fields."}, status=400)

        case_data = Case.objects.filter(case_id=case_id).first()
        if case_data.case_type == 'dc_visit':
            case_type = 'dc_visit'
        elif case_data.case_type == 'both':
            case_type = case_data.case_stage
        else:
            case_type = 'vmer'

        Schedule.objects.filter(case_id=case_id, is_active=True).update(is_active=False)
        already_scheduled = Schedule.objects.filter(case_id=case_id, schedule_type=case_type).exists()

        schedule = Schedule.objects.create(
            case_id=case_id,
            schedule_time=schedule_time,
            schedule_type=case_type,
            reason=reason,
            created_by=request.user.user_id
        )

        dt = datetime.strptime(str(schedule_time), "%Y-%m-%d %H:%M:%S")
        date = dt.strftime("%d-%b-%Y")
        time = dt.strftime("%I:%M %p")

        if already_scheduled:
            Case.objects.filter(case_id=case_id).update(status='rescheduled')
            CaseActionLog.objects.create(case_id=case_id, action_by=request.user.user_id, action="ReSchedule Created")

        else:
            Case.objects.filter(case_id=case_id).update(status='scheduled')
            CaseActionLog.objects.create(case_id=case_id, action_by=request.user.user_id, action="Schedule Created")

        dc_data = DiagnosticCenter.objects.filter(user_id=case_data.assigned_dc_id).first()
        dc_user_data = User.objects.filter(user_id=case_data.assigned_dc_id).first()
        try:
            send_scheduled(
                date=date,
                time=time,
                dc_name=dc_data.name,
                address=dc_data.address,
                gmap_link=f"https://www.google.com/maps/search/{str(dc_data.name).replace(" ", "%20")}+{str(dc_data.address).replace(" ", "%20")}+{str(dc_data.city).replace(" ", "%20")}+{str(dc_data.state).replace(" ", "%20")}",
                contact_number="7718861051",
                email_id="lic.pims@ericsontpa.com",
                recipient_email=case_data.holder_email,
                phone=case_data.holder_phone,
            )
            time_fun.sleep(5)
            send_medical_email(                    
                    recipient_email=dc_user_data.email,
                    appointment_date=date,
                    subject=f"Appointment on {date} {case_data.holder_name}",
                    insurance_company="LIC OF INDIA",
                    intimation_number=case_id,
                    branch_code=case_data.ins_office_code,
                    proposal_number=case_data.policy_number,
                    client_name=case_data.holder_name,
                    dob=f"{case_data.holder_dob.day}/{case_data.holder_dob.month}/{case_data.holder_dob.year}",
                    gender=str(case_data.holder_gender).upper(),
                    contact_number=case_data.holder_phone,
                    sum_assured=case_data.sum_insured_under_consideration,
                    medical_test="#".join(case_data.tests),
                    intimation_date=f"{case_data.intimation_date.day}/{case_data.intimation_date.month}/{case_data.intimation_date.year}",
                    appointment_time=time,
                    visit_type="Centre visit",
                    client_address=case_data.holder_address
                )
        except:
            pass
        return Response({"success": True, "data": ScheduleSerializer(schedule).data})


class CaseIssueViewSet(viewsets.ViewSet):

    @check_authentication(required_role=['vmer_med_co', 'diagnostic_center'])
    @handle_exceptions
    def create(self, request):
        case_id = request.data.get('case_id')
        issue_type = request.data.get('issue_type')
        reason = request.data.get('reason')

        if not all([case_id, issue_type, reason]):
            return Response({"error": "Missing required fields."}, status=400)

        case_obj = Case.objects.filter(case_id=case_id, is_active=True).first()
        if not case_obj:
            return Response({"error": "Invalid case_id."}, status=404)

        case_obj.issue_type = issue_type
        case_obj.issue_reason = reason
        case_obj.status = 'issue'
        case_obj.save()

        # Log the action
        action_text = f"Issue reported: {str(issue_type).replace('_', ' ').title()}"
        CaseActionLog.objects.create(
            case_id=case_id,
            action_by=request.user.user_id,
            action=action_text,
            remarks=reason
        )

        return Response({
            "success": True,
            "message": "Issue reported successfully. Case will be rescheduled by telecaller.",
            "data": CaseSerializer(case_obj).data
        }, status=201)


class CaseLogViewSet(viewsets.ViewSet):

    @check_authentication()
    @handle_exceptions
    def list(self, request):
        case_id = request.query_params.get("case_id")
        if not case_id:
            return Response({"error": "Missing case_id."}, status=400)

        logs = CaseActionLog.objects.filter(case_id=case_id).order_by('-timestamp')
        serializer = CaseActionLogSerializer(logs, many=True)
        return Response({"success": True, "data": serializer.data})


class StaffListViewSet(viewsets.ViewSet):

    @check_authentication()
    @handle_exceptions
    def list(self, request):
        role = request.query_params.get('role')
        if not role:
            return Response({"error": "Missing role param."}, status=400)

        users = User.objects.filter(role=role, is_active=True)
        serializer = UserSerializer(users, many=True)
        return Response({"success": True, "data": serializer.data})


class UploadDocumentViewSet_old(viewsets.ViewSet):

    @check_authentication(required_role='vmer_med_co')
    @handle_exceptions
    def create(self, request):
        case_id = request.data.get('case_id')
        video_url = request.data.get('video_url')

        if not case_id or not video_url:
            return Response({"error": "Missing case_id or video_url."}, status=400)

        case = Case.objects.filter(case_id=case_id, is_active=True).first()
        if not case:
            return Response({"error": "Invalid case_id."}, status=404)

        case.video_url = video_url
        case.status = 'uploaded'
        case.save()

        CaseActionLog.objects.create(
            case_id=case_id,
            action_by=request.user.user_id,
            action="Video recording uploaded by VMER Med Co"
        )

        return Response({"success": True, "message": "Video uploaded and submitted to LIC."}, status=200)

    @check_authentication(required_role='diagnostic_center')
    @handle_exceptions
    def update(self, request, pk=None):
        case_id = request.data.get('case_id')
        report_url = request.data.get('report_url')

        if not case_id or not report_url:
            return Response({"error": "Missing case_id or report_url."}, status=400)

        case = Case.objects.filter(case_id=case_id, is_active=True).first()
        if not case:
            return Response({"error": "Invalid case_id."}, status=404)

        case.report_url = report_url
        case.status = 'uploaded'
        case.save()

        CaseActionLog.objects.create(
            case_id=case_id,
            action_by=request.user.user_id,
            action="Diagnostic report uploaded by DC"
        )

        return Response({"success": True, "message": "Report uploaded and submitted to LIC."}, status=200)


class UploadDocumentViewSet(viewsets.ViewSet):

    @check_authentication(required_role=['vmer_med_co', 'diagnostic_center'])
    @handle_exceptions
    def create(self, request):
        case_id = request.data.get('case_id')
        uploaded_file = request.FILES.get('file')

        if not case_id or not uploaded_file:
            return Response({"error": "Missing case_id or file."}, status=400)

        case = Case.objects.get(case_id=case_id, is_active=True)
        if not case:
            return Response({"error": "Invalid case_id."}, status=404)

        try:
            # Upload file to S3
            file_url = upload_file_to_s3(uploaded_file)

            if request.user.role == 'vmer_med_co':
                case.video_url = file_url
                action = "Video recording uploaded by VMER Med Co"
            elif request.user.role == 'diagnostic_center':
                case.report_url = file_url
                action = "Diagnostic report uploaded by DC"

            case.status = 'uploaded'
            if str(case.case_type).lower() == "both":
                if case.case_stage == "vmer":
                    case.case_stage == "dc_visit"
                    case.status = 'assigned'
                    case.issue_reason = ''
                    case.issue_type = ''

            case.save()

            CaseActionLog.objects.create(
                case_id=case_id,
                action_by=request.user.user_id,
                action=action
            )

            return Response({"success": True, "message": "File uploaded successfully.", "file_url": file_url}, status=200)

        except Exception as e:
            return Response({"error": f"File upload failed: {str(e)}"}, status=500)


class DiagnosticCenterViewSet(viewsets.ViewSet):

    @check_authentication(required_role='admin')  # or 'hod' if needed
    @handle_exceptions
    def create(self, request):
        name = request.data.get("name")
        email = request.data.get("email")
        password = request.data.get("password")
        address = request.data.get("address")
        city = request.data.get("city")
        state = request.data.get("state")
        pincode = request.data.get("pincode")
        contact_person = request.data.get("contact_person")
        contact_number = request.data.get("contact_number")

        if not all([name, email, password, address, city, state, pincode]):
            return Response({
                "success": False,
                "data": None,
                "error": "Missing required fields."
            }, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email=email, is_active=True).exists():
            return Response({
                "success": False,
                "data": None,
                "error": "A user with this email already exists."
            }, status=status.HTTP_400_BAD_REQUEST)

        user_id = self.generate_unique_user_id()

        # Create user
        user = User.objects.create_user(
            user_id=user_id,
            username=user_id,
            password=password,
            email=email,
            name=name,
            contact_number=contact_number,
            role="diagnostic_center"
        )

        # Create Diagnostic Center entry
        dc = DiagnosticCenter.objects.create(
            user_id=user_id,
            name=name,
            address=address,
            city=city,
            state=state,
            pincode=pincode,
            contact_person=contact_person,
            contact_number=contact_number
        )

        return Response({
            "success": True,
            "data": {
                "user_id": user_id,
                "email": email,
                "diagnostic_center": DiagnosticCenterSerializer(dc).data
            },
            "error": None
        }, status=status.HTTP_201_CREATED)

    def generate_unique_user_id(self):
        prefix = "DC"
        while True:
            random_id = ''.join(random.choices(string.digits, k=10))
            user_id = prefix + random_id
            if not User.objects.filter(user_id=user_id).exists():
                return user_id


class ReportDownloadViewSet(viewsets.ViewSet):

    @check_authentication(required_role='admin')
    @handle_exceptions
    def list(self, request):
        return Response({"message": "Use POST methods to download invoices."})

    @check_authentication(required_role='admin')
    @handle_exceptions
    def create(self, request):
        report_type = request.data.get("report_type")
        if report_type == "dc_invoice":
            return self.generate_dc_invoice(request)
        elif report_type == "lic_invoice":
            return self.generate_lic_invoice(request)
        elif report_type == "coordinator_report":
            return self.generate_coordinator_report(request)
        else:
            return Response({"error": "Invalid report_type."}, status=400)

    def generate_dc_invoice(self, request):
        dc_user_id = request.data.get("dc_user_id")
        month = request.data.get("month")  # format: YYYY-MM

        year, month_num = map(int, month.split('-'))
        start_date = make_aware(datetime(year, month_num, 1))
        end_day = calendar.monthrange(year, month_num)[1]
        end_date = make_aware(datetime(year, month_num, end_day, 23, 59, 59))

        cases = Case.objects.filter(
            assigned_dc_id=dc_user_id,
            status='submitted_to_lic',
            created_at__range=(start_date, end_date)
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "DC Invoice"
        ws.append(["Case ID", "Holder Name", "Phone", "Case Type", "Completed At", "Amount"])

        for case in cases:
            ws.append([
                case.case_id,
                case.holder_name,
                case.holder_phone,
                case.case_type,
                case.updated_at.strftime('%Y-%m-%d %H:%M'),
                "--"  # Amount can be fetched from a config if added later
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=dc_invoice_{month}.xlsx'
        wb.save(response)
        return response

    def generate_lic_invoice(self, request):
        ins_office_code = request.data.get("ins_office_code")
        month = request.data.get("month")

        year, month_num = map(int, month.split('-'))
        start_date = make_aware(datetime(year, month_num, 1))
        end_day = calendar.monthrange(year, month_num)[1]
        end_date = make_aware(datetime(year, month_num, end_day, 23, 59, 59))

        cases = Case.objects.filter(
            ins_office_code=ins_office_code,
            payment_method='lic',
            status='submitted_to_lic',
            created_at__range=(start_date, end_date)
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "LIC Invoice"
        ws.append(["Case ID", "Holder Name", "Policy No", "Case Type", "Completed At", "Amount"])

        for case in cases:
            ws.append([
                case.case_id,
                case.holder_name,
                case.policy_number,
                case.case_type,
                case.updated_at.strftime('%Y-%m-%d %H:%M'),
                "--"
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=lic_invoice_{month}.xlsx'
        wb.save(response)
        return response

    def generate_coordinator_report(self, request):
        coordinator_id = request.data.get("coordinator_id")
        month = request.data.get("month")

        year, month_num = map(int, month.split('-'))
        start_date = make_aware(datetime(year, month_num, 1))
        end_day = calendar.monthrange(year, month_num)[1]
        end_date = make_aware(datetime(year, month_num, end_day, 23, 59, 59))

        cases = Case.objects.filter(
            assigned_coordinator_id=coordinator_id,
            created_at__range=(start_date, end_date)
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Coordinator Report"
        ws.append(["Case ID", "Holder Name", "Case Type", "Status", "Created At"])

        for case in cases:
            ws.append([
                case.case_id,
                case.holder_name,
                case.case_type,
                case.status,
                case.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=coordinator_report_{month}.xlsx'
        wb.save(response)
        return response


class GetTestDetailsViewSet(viewsets.ViewSet):

    @handle_exceptions
    def list(self, request):
        test_details = TestDetail.objects.all()
        test_details_data = TestDetailSerializer(test_details, many=True).data
        return Response({
            "success": True,
            "data": test_details_data
        })

class TelecallerRemarkViewSet(viewsets.ViewSet):

    @check_authentication(required_role=['telecaller'])
    @handle_exceptions
    def create(self, request):
        case_id = request.data.get("case_id")
        remark = request.data.get("remark")

        if not case_id or not remark:
            return Response({"success": False, "error": "case_id and remark are required."}, status=400)

        case = Case.objects.filter(case_id=case_id, is_active=True).first()
        if not case:
            return Response({"success": False, "error": "Invalid case_id"}, status=404)

        telecaller_remark = TelecallerRemark.objects.create(
            case_id=case_id,
            telecaller_id=request.user.user_id,
            remark=remark
        )
        CaseActionLog.objects.create(case_id=case_id, action_by=request.user.user_id, action="Telecaller added call remark")
        return Response({"success": True, "data": TelecallerRemarkSerializer(telecaller_remark).data}, status=201)

    @check_authentication(required_role=['telecaller', 'coordinator', 'hod', 'admin'])
    @handle_exceptions
    def list(self, request):
        case_id = request.query_params.get("case_id")
        if not case_id:
            return Response({"success": False, "error": "Missing case_id"}, status=400)

        remarks = TelecallerRemark.objects.filter(case_id=case_id)
        serializer = TelecallerRemarkSerializer(remarks, many=True)
        return Response({"success": True, "data": serializer.data})


def get_date_filter(month=None, year=None, fy=None, start_date=None, end_date=None):
    """Reusable date filter builder"""
    if month and year:
        year, month_num = int(year), int(month)
        start = make_aware(datetime(year, month_num, 1))
        end = make_aware(datetime(year, month_num, calendar.monthrange(year, month_num)[1], 23, 59, 59))
        return {"created_at__range": (start, end)}
    elif year:
        start = make_aware(datetime(int(year), 1, 1))
        end = make_aware(datetime(int(year), 12, 31, 23, 59, 59))
        return {"created_at__range": (start, end)}
    elif fy:
        start_year, end_year = map(int, fy.split("-"))
        start = make_aware(datetime(start_year, 4, 1))
        end = make_aware(datetime(end_year, 3, 31, 23, 59, 59))
        return {"created_at__range": (start, end)}
    elif start_date and end_date:
        start = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
        end = make_aware(datetime.strptime(end_date, "%Y-%m-%d"))
        return {"created_at__range": (start, end)}
    return {}


# --------------------------
# REPORTS
# --------------------------
class ReportSummaryViewSet(viewsets.ViewSet):
    """
    Case Reports by Branch → Division → Region → Head Office
    """

    @check_authentication(required_role=['admin', 'hod'])
    @handle_exceptions
    def list(self, request):
        # Filters
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        fy = request.query_params.get("fy")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        date_filter = get_date_filter(month, year, fy, start_date, end_date)

        all_cases = Case.objects.filter(is_active=True, **date_filter)

        branch_data, division_data, region_data, ho_data = {}, {}, {}, {}

        for case in all_cases:
            branch_id = case.ins_office_code
            if not branch_id:
                continue

            branch = BranchOffice.objects.filter(name=branch_id).first()
            if not branch:
                continue

            # ---- Branch Level ----
            if branch_id not in branch_data:
                branch_data[branch_id] = {"cases": 0, "completed": 0, "pending": 0}
            branch_data[branch_id]["cases"] += 1
            if case.status == "completed":
                branch_data[branch_id]["completed"] += 1
            else:
                branch_data[branch_id]["pending"] += 1

            # ---- Division Level ----
            division_id = branch.divisional_office_id
            if division_id not in division_data:
                division_data[division_id] = {"cases": 0}
            division_data[division_id]["cases"] += 1

            # ---- Region Level ----
            division = DivisionalOffice.objects.filter(lic_id=division_id).first()
            if division:
                region_id = division.regional_office_id
                if region_id not in region_data:
                    region_data[region_id] = {"cases": 0}
                region_data[region_id]["cases"] += 1

                # ---- Head Office ----
                region = RegionalOffice.objects.filter(lic_id=region_id).first()
                if region:
                    ho_id = region.head_office_id
                    if ho_id not in ho_data:
                        ho_data[ho_id] = {"cases": 0}
                    ho_data[ho_id]["cases"] += 1

        return Response({
            "success": True,
            "data": {
                "branch": branch_data,
                "division": division_data,
                "region": region_data,
                "head_office": ho_data
            }
        }, status=status.HTTP_200_OK)


# --------------------------
# FINANCE - LIC
# --------------------------
class FinanceLICViewSet(viewsets.ViewSet):
    """
    Finance: Money to Collect from LIC (Branch → Division → Region → HO)
    """

    @check_authentication(required_role=['admin', 'hod', 'accounts'])
    @handle_exceptions
    def list(self, request):
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        fy = request.query_params.get("fy")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        date_filter = get_date_filter(month, year, fy, start_date, end_date)

        cases = Case.objects.filter(payment_method="lic", is_active=True, **date_filter)

        branch_data, division_data, region_data, ho_data = {}, {}, {}, {}

        for case in cases:
            branch_id = case.ins_office_code
            if not branch_id:
                continue

            branch = BranchOffice.objects.filter(name=branch_id).first()
            if not branch:
                continue

            # Case amount
            total_case_amount = 0
            for _, prices in case.test_price.items():
                if case.lic_type == "urban":
                    total_case_amount += int(prices.get("lic_urban_charge", 0))
                else:
                    total_case_amount += int(prices.get("lic_rural_charge", 0))

            # ---- Branch ----
            if branch_id not in branch_data:
                branch_data[branch_id] = {"cases": 0, "total_amount": 0}
            branch_data[branch_id]["cases"] += 1
            branch_data[branch_id]["total_amount"] += total_case_amount

            # ---- Division ----
            division_id = branch.divisional_office_id
            if division_id not in division_data:
                division_data[division_id] = {"cases": 0, "total_amount": 0}
            division_data[division_id]["cases"] += 1
            division_data[division_id]["total_amount"] += total_case_amount

            # ---- Region ----
            division = DivisionalOffice.objects.filter(lic_id=division_id).first()
            if division:
                region_id = division.regional_office_id
                if region_id not in region_data:
                    region_data[region_id] = {"cases": 0, "total_amount": 0}
                region_data[region_id]["cases"] += 1
                region_data[region_id]["total_amount"] += total_case_amount

                # ---- HO ----
                region = RegionalOffice.objects.filter(lic_id=region_id).first()
                if region:
                    ho_id = region.head_office_id
                    if ho_id not in ho_data:
                        ho_data[ho_id] = {"cases": 0, "total_amount": 0}
                    ho_data[ho_id]["cases"] += 1
                    ho_data[ho_id]["total_amount"] += total_case_amount

        return Response({
            "success": True,
            "data": {
                "branch": branch_data,
                "division": division_data,
                "region": region_data,
                "head_office": ho_data
            }
        }, status=status.HTTP_200_OK)


# --------------------------
# FINANCE - DC
# --------------------------
class FinanceDCViewSet(viewsets.ViewSet):
    """
    Finance: Payouts to DCs
    """

    @check_authentication(required_role=['admin', 'hod', 'accounts'])
    @handle_exceptions
    def list(self, request):
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        fy = request.query_params.get("fy")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        date_filter = get_date_filter(month, year, fy, start_date, end_date)

        cases = Case.objects.filter(case_type__in=["dc_visit", "both"], is_active=True, **date_filter)

        dc_data = {}
        for case in cases:
            dc_id = case.assigned_dc_id
            if not dc_id:
                continue

            dc_user = User.objects.filter(user_id=dc_id).first()
            dc_name = dc_user.name if dc_user else "Unknown DC"

            if dc_id not in dc_data:
                dc_data[dc_id] = {"dc_name": dc_name, "cases": 0, "total_amount": 0}

            total_case_amount = 0
            for _, prices in case.test_price.items():
                total_case_amount += int(prices.get("dc_charge", 0))

            dc_data[dc_id]["cases"] += 1
            dc_data[dc_id]["total_amount"] += total_case_amount

        return Response({"success": True, "data": dc_data}, status=status.HTTP_200_OK)


def AddTests(request):
    error = []
    testss = [
        ["TMG - JMR", 28, 40],
        ["TMG - Urine Routine", 28, 40],
        ["TMG - Urine Routine, BPB", 287, 410],
        ["TMG - Urine Routine, BPB, HBa1c", 336, 480],
        ["TMG - Urine Routine, BPB, HBa1c, ECG", 532, 760],
        ["TMG - Urine Routine, BPB, CBC, ECG", 532, 760],
        ["TMG - Urine Routine, BPB,CBC, Hba1c", 385, 550],
        ["TMG - Urine Routine, BPB,CBC, HBA1c, ECG", 581, 830],
        ["TMG - Urine Routine, BPB - F, CBC, HBA1c, ECG", 581, 830],
        ["TMG - Urine Routine, Cotinine", 91, 130],
        ["TMG - Urine Routine, BPB, Cotinine", 365, 522],
        ["TMG - Urine Routine, BPB, HBa1c, Cotinine", 434, 620],
        ["TMG - Urine Routine, BPB, HBa1c, ECG, Cotinine", 630, 900],
        ["TMG - Urine Routine, BPB, CBC, ECG, Cotinine", 630, 900],
        ["TMG - Urine Routine, BPB,CBC, Hba1c, Cotinine", 483, 690],
        ["TMG - Urine Routine, BPB,CBC, HBA1c, ECG, Cotinine", 679, 970],
        ["TMG - Urine Routine, BPB - F, CBC, HBA1c, ECG, Cotinine", 679, 970],
        ["TMG - Urine Routine, BPB -F,CBC,HBA1c", 413, 590],
        ["TMG - Videography MER", 98, 140],
        ["TMG - Videography MER, Urine Routine", 98, 140],
        ["TMG - Videography MER, Urine Routine, BPB", 392, 560],
        ["TMG - Videography MER, Urine Routine, BPB, HBa1c", 441, 630],
        ["TMG - Videography MER, Urine Routine, BPB, HBa1c, ECG", 637, 910],
        ["TMG - Videography MER, Urine Routine, BPB, CBC, ECG", 637, 910],
        ["TMG - Videography MER, Urine Routine, BPB,CBC, Hba1c", 490, 700],
        ["TMG - Videography MER, Urine Routine, BPB,CBC, HBA1c, ECG", 686, 980],
        ["TMG - Videography MER, Urine Routine, BPB - F, CBC, HBA1c, ECG", 686, 980],
        ["TMG - Videography MER, Urine Routine, Cotinine", 196, 280],
        ["TMG - Videography MER, Urine Routine, BPB, Cotinine", 470, 672],
        ["TMG - Videography MER, Urine Routine, BPB, HBa1c, Cotinine", 539, 770],
        ["TMG - Videography MER, Urine Routine, BPB, HBa1c, ECG, Cotinine", 735, 1050],
        ["TMG - Videography MER, Urine Routine, BPB, CBC, ECG, Cotinine", 735, 1050],
        ["TMG - Videography MER, Urine Routine, BPB,CBC, Hba1c, Cotinine", 588, 840],
        ["TMG - Videography MER, Urine Routine, BPB,CBC, HBA1c, ECG, Cotinine", 784, 1120],
        ["TMG - Videography MER, Urine Routine, BPB - F, CBC, HBA1c, ECG, Cotinine", 784, 1120],
        ["TMG - Videography MER, Urine Routine, BPB -F,CBC,HBA1c", 518, 740],
        ["Ericson - Microscopic Urine Analysis", 32, 45],
        ["Ericson - ECG-Resting (ECG)", 84, 120],
        ["Ericson - Anti HCV (IgG) (HCV Antibodies)", 210, 300],
        ["Ericson - APTT (Activated Partial Thromboplastin Time)", 140, 200],
        ["Ericson - Bilirubin (Direct/Conjugated)", 28, 40],
        ["Ericson - Bilirubin (Indirect/Unconjugated)", 28, 40],
        ["Ericson - Bilirubin (Total)", 28, 40],
        ["Ericson - Calcium", 28, 40],
        ["Ericson - ESR (Erythrocyte Sedimentation rate)", 28, 40],
        ["Ericson - FSH(Follicle Stimulating Hormone)", 70, 100],
        ["Ericson - Haemogram (without peripheral smear)", 140, 200],
        ["Ericson - Hb% (Haemoglobin)", 28, 40],
        ["Ericson - Random Blood Sugar (RBS)", 28, 40],
        ["Ericson - Fasting Blood Sugar (FBS)", 28, 40],
        ["Ericson - Alkaline Phosphates", 28, 40],
        ["Ericson - Gamma GT", 28, 40],
        ["Ericson - HbA1c (Glycosylated Haemoglobin)", 105, 150],
        ["Ericson - HbeAg", 175, 250],
        ["Ericson - HCV(Hepatitis C Virus)", 210, 300],
        ["Ericson - HDL(High Density Lipids Cholesterol)", 28, 40],
        ["Ericson - MCH (Mean Corpuscular Haemoglobin)", 28, 40],
        ["Ericson - MCH (Mean Corpuscular Haemoglobin Concentration)", 28, 40],
        ["Ericson - MCV (Mean Corpuscular Volume)", 28, 40],
        ["Ericson - Mg(Magnesium)", 70, 100],
        ["Ericson - PCV(Haematocrit)", 28, 40],
        ["Ericson - Peripheral Blood Smear (without blood profile)", 84, 120],
        ["Ericson - Platelets", 28, 40],
        ["Ericson - Proteins(Globulin)", 28, 40],
        ["Ericson - Proteins(Total)", 28, 40],
        ["Ericson - PSA(Prostate Specific Antigen)", 175, 250],
        ["Ericson - PTT(Partial Thromboplastin Time)", 140, 200],
        ["Ericson - RBC (Red Blood Cells)", 28, 40],
        ["Ericson - Serum Acid Phosphates", 105, 150],
        ["Ericson - Serum Amylase", 280, 400],
        ["Ericson - Serum CPK (Creatinine Phosphokinase)", 105, 150],
        ["Ericson - Serum LH (Luetenizing Hormone)", 70, 100],
        ["Ericson - Serum Lipase", 350, 500],
        ["Ericson - Serum Uric Acid", 28, 40],
        ["Ericson - Sodium", 28, 40],
        ["Ericson - Stool Routine", 28, 40],
        ["Ericson - T3(Free)", 28, 40],
        ["Ericson - T3(Triiodothyronine)", 28, 40],
        ["Ericson - T4(Free)", 28, 40],
        ["Ericson - T4 (Thyroxine)", 28, 40],
        ["Ericson - Total Leucocyte Count(TLC)", 28, 40],
        ["Ericson - Total Lipids", 91, 130],
        ["Ericson - TSH (Thyroid Stimulating Hormone)", 140, 200],
        ["Ericson - VLDL(Very Low Density Lipids)", 28, 40],
        ["Ericson - WBC(White Blood Cells)", 28, 40],
        ["Ericson - Lipid Profile (Total Cholesterol, LDL Cholesterol, HDL Cholesterol, Triglycerides, VLDL Cholesterol)", 91, 130],
        ["Ericson - Diabetic Profile (FBS, PPBS & HBA1C)", 168, 240],
        ["Ericson - Liver Function Test (ALT, AST, ALP, Total Bilirubin, Albumin, Total Protein, GGT, Bilirubin Direct, A/G Ratio)", 91, 130],
        ["Ericson - Thyroid Profile (T3 (Triiodothyronine), T4 (Thyroxine), TSH (Thyroid Stimulating Hormone))", 140, 200],
        ["Ericson - CBC (RBCs, WBCs, Platelet Count, Mean Platelet Volume)", 105, 150],
        ["Ericson - CBC with ESR (RBCs, WBCs, Platelet Count, Mean Platelet Volume)", 133, 190],
        ["Ericson - HIV", 140, 200],
        ["Ericson - Cholesterol", 28, 40],
        ["Ericson - HDL Cholesterol", 28, 40],
        ["Ericson - Triglycerides", 28, 40],
        ["Ericson - Creatinine", 28, 40],
        ["Ericson - Urea(BUN)", 28, 40],
        ["Ericson - HbsAg", 140, 200],
        ["Ericson - SGOT", 28, 40],
        ["Ericson - SGPT", 28, 40],
        ["Ericson - Alk Phosphates", 28, 40],
        ["Ericson - GGTP", 28, 40],
        ["Ericson - Albumin", 28, 40],
        ["Ericson - Serum Albumin", 28, 40],
        ["Ericson - FBS", 28, 40],
        ["Ericson - PG2", 140, 200],
        ["Ericson - 2 Urine dipstick", 28, 40],
        ["Ericson - 3BP Reading", 11, 15],
        ["Ericson - BPB (HIV, Cholesterol, Triglycerides, Random Blood Sugar, Serum Creatinine, HDL, S.Bilirubin, HbsAg, SGOT, SGPT, Alkaline Phosphates, Gamma GT, Serum Albumin, BUN)", 392, 560],
        ["Ericson - BPB-F (FBS-Fasting Blood Sugar, HIV, Cholesterol, Triglycerides, Serum Creatinine, HDL, Serum Bilirubin, HbsAg, SGOT, SGPT, Alkaline Phosphates, Gamma GT, Serum Albumin, BUN)", 392, 560],
        ["Ericson - Urine Cotinine", 140, 200],
        ["Ericson - Routine Urine Analysis (RUA)", 28, 40],
        ["Ericson - Microalbumin(Urine)", 70, 100],
        ["Ericson - Renal / Kidney Function Test (Serum Creatinine, BUN, GFR, Electrolytes, Uric Acid, Albumin-to-Creatinine Ratio)", 91, 130],
        ["ECG", 84, 120],
        ["Ericson - Vision Testing", 70, 100],
        ["Ericson - X-Ray (Barium Enema)", "As per acutal", "As per acutal"],
        ["Ericson - X-Ray (Barium Swallow)", "As per acutal", "As per acutal"],
        ["Ericson - X-ray (IVP-Intravenous Pylegraphy)", "As per acutal", "As per acutal"],
        ["Ericson - X-ray (Lumbo-Sacral)", "As per acutal", "As per acutal"],
        ["Ericson - Stress Echo", 595, 850],
        ["Ericson - 2D ECHO", 560, 800],
        ["Ericson - Fundoscopy", 175, 250],
        ["Ericson - Stress Thalium", "As per acutal", "As per acutal"],
        ["Ericson - Chest X-Ray", 84, 120],
        ["Ericson - X-ray KUB", 91, 130],
        ["Ericson - USG - Whole Abdomen", 490, 700],
        ["Ericson - USG - KUB", 490, 700],
        ["Ericson - USG - Thyroid", 490, 700],
        ["Ericson - USG - Pelvis", 490, 700],
        ["Ericson - USG - KUB & Abdomen", 525, 750],
        ["Ericson - USG - Pelvis & Abdomen", 525, 750],
        ["Ericson - PFT", 105, 150],
        ["Ericson - TMT(Tread Mill Test)", 525, 750],
        ["Ericson - GTT(Glucose Tolerance Test-5 samples)", 28, "40 per sample"],
        ["Ericson - BST(Blood Sugar Tolerance)", 28, 40],
        ["Ericson - BT(Bleeding Time)", 28, 40],
        ["Ericson - HBDH(Hydroxy Butyrate DeHydrogenase)", "As per acutal", "As per acutal"],
        ["Ericson - CEA (Carcinoembryonic Antigen)", 140, 200],
        ["Ericson - Alpha-Fetoprotein (AFP)", 140, 200],
        ["Ericson - Audiometry", 70, 100],
        ["Ericson - Vitamin B12", 105, 150],
        ["Ericson - Vitamin D (25-OH)", 105, 150],
        ["Ericson - Vitamin B 9 (Folate)", 105, 150],
        ["Ericson - General Physician / Doctor Consultation", 140, 200],
        ["Ericson - Gynec Consultation", 280, 400],
        ["Ericson - PAP Smear", 175, 250],
        ["Physical MER at Center Visit", 70, 100],

    ]
    for testt in testss:
        if TestDetail.objects.filter(test_name=testt[0]).exists():
            print("Test already exists, skipping:", testt[0])
            error.append({"test": testt[0], "error": "Test already exists, skipping"})
            continue
        if type(testt[1]) != int or type(testt[2]) != int:
            print("Skipping invalid test entry:", testt)
            error.append({"test": testt[0], "error": "Skipping invalid test entry"})
            continue
        TestDetail.objects.create(
            test_name=testt[0],
            dc_charge=testt[1],
            lic_rural_charge=testt[2],
            lic_urban_charge=testt[2],
            is_active=True
        )
    return JsonResponse({"success": True, "data": error}, status=200)

# def AddTests(request):
#     test_a = TestDetail.objects.all()
#     for test in test_a:
#         print(test.test_name)
#     return JsonResponse({"success": True, "data": "Tests listed"}, status=200)

class InsuranceCompanyViewSet(viewsets.ViewSet):
    """
    ViewSet for managing insurance companies
    """
    
    @check_authentication(required_role=['admin', 'hod'])
    @handle_exceptions
    def create(self, request):
        serializer = InsuranceCompanySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True, "data": serializer.data, "error": None}, status=status.HTTP_201_CREATED)
        return Response({"success": False, "data": None, "error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @check_authentication()
    @handle_exceptions
    def list(self, request):
        companies = InsuranceCompany.objects.filter(is_active=True)
        serializer = InsuranceCompanySerializer(companies, many=True)
        return Response({"success": True, "data": serializer.data, "error": None})

    @check_authentication(required_role=['admin', 'hod'])
    @handle_exceptions
    def update(self, request, pk=None):
        company = InsuranceCompany.objects.filter(id=pk).first()
        if not company:
            return Response({"success": False, "data": None, "error": "Not found."}, status=404)

        serializer = InsuranceCompanySerializer(company, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True, "data": serializer.data, "error": None})
        return Response({"success": False, "data": None, "error": serializer.errors}, status=400)


class TataAIGOfficeViewSet(viewsets.ViewSet):
    """
    ViewSet for managing Tata AIG offices (simple structure, no hierarchy)
    """
    
    @check_authentication(required_role=['admin', 'hod'])
    @handle_exceptions
    def create(self, request):
        serializer = TataAIGOfficeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True, "data": serializer.data, "error": None}, status=status.HTTP_201_CREATED)
        return Response({"success": False, "data": None, "error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @check_authentication()
    @handle_exceptions
    def list(self, request):
        offices = TataAIGOffice.objects.filter(is_active=True)
        serializer = TataAIGOfficeSerializer(offices, many=True)
        return Response({"success": True, "data": serializer.data, "error": None})

    @check_authentication(required_role=['admin', 'hod'])
    @handle_exceptions
    def update(self, request, pk=None):
        office = TataAIGOffice.objects.filter(id=pk).first()
        if not office:
            return Response({"success": False, "data": None, "error": "Not found."}, status=404)

        serializer = TataAIGOfficeSerializer(office, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True, "data": serializer.data, "error": None})
        return Response({"success": False, "data": None, "error": serializer.errors}, status=400)

    @check_authentication(required_role=['admin', 'hod'])
    @handle_exceptions
    def destroy(self, request, pk=None):
        office = TataAIGOffice.objects.filter(id=pk).first()
        if not office:
            return Response({"success": False, "data": None, "error": "Not found."}, status=404)
        office.is_active = False
        office.save()
        return Response({"success": True, "data": {"id": pk, "deleted": True}, "error": None})


class FinanceInsuranceViewSet(viewsets.ViewSet):
    """
    Finance: Money to Collect from Insurance Companies (supports both LIC hierarchy and simple offices)
    Returns ALL financial data - filtering is handled on frontend
    """

    @check_authentication(required_role=['admin', 'hod', 'accounts'])
    @handle_exceptions
    def list(self, request):
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        fy = request.query_params.get("fy")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        insurance_company_id = request.query_params.get("insurance_company_id")

        if not insurance_company_id:
            return Response({"success": False, "error": "insurance_company_id is required"}, status=400)

        # Get insurance company
        insurance_company = InsuranceCompany.objects.filter(id=insurance_company_id, is_active=True).first()
        if not insurance_company:
            return Response({"success": False, "error": "Invalid insurance company"}, status=404)

        date_filter = get_date_filter(month, year, fy, start_date, end_date)

        # Filter cases by insurance company
        cases = Case.objects.filter(
            insurance_company=insurance_company.code,
            payment_method="lic",  # This should be renamed to "insurance" in future
            is_active=True,
            **date_filter
        )

        if insurance_company.has_hierarchy:
            # LIC-style hierarchy
            return self._process_hierarchical_offices(cases, insurance_company)
        else:
            # Simple office structure (Tata AIG, etc.)
            return self._process_simple_offices(cases, insurance_company)

    def _process_hierarchical_offices(self, cases, insurance_company):
        """Process cases for insurance companies with office hierarchy (like LIC)"""
        branch_data, division_data, region_data, ho_data = {}, {}, {}, {}

        for case in cases:
            branch_id = case.ins_office_code
            if not branch_id:
                continue

            branch = BranchOffice.objects.filter(name=branch_id).first()
            if not branch:
                continue

            # Calculate case amount
            total_case_amount = 0
            for _, prices in case.test_price.items():
                if case.lic_type == "urban":
                    total_case_amount += int(prices.get("lic_urban_charge", 0))
                else:
                    total_case_amount += int(prices.get("lic_rural_charge", 0))

            # Branch level
            if branch_id not in branch_data:
                branch_data[branch_id] = {"cases": 0, "total_amount": 0}
            branch_data[branch_id]["cases"] += 1
            branch_data[branch_id]["total_amount"] += total_case_amount

            # Division level
            division_id = branch.divisional_office_id
            if division_id not in division_data:
                division_data[division_id] = {"cases": 0, "total_amount": 0}
            division_data[division_id]["cases"] += 1
            division_data[division_id]["total_amount"] += total_case_amount

            # Region level
            division = DivisionalOffice.objects.filter(lic_id=division_id).first()
            if division:
                region_id = division.regional_office_id
                if region_id not in region_data:
                    region_data[region_id] = {"cases": 0, "total_amount": 0}
                region_data[region_id]["cases"] += 1
                region_data[region_id]["total_amount"] += total_case_amount

                # Head Office level
                region = RegionalOffice.objects.filter(lic_id=region_id).first()
                if region:
                    ho_id = region.head_office_id
                    if ho_id not in ho_data:
                        ho_data[ho_id] = {"cases": 0, "total_amount": 0}
                    ho_data[ho_id]["cases"] += 1
                    ho_data[ho_id]["total_amount"] += total_case_amount

        return Response({
            "success": True,
            "data": {
                "branch": branch_data,
                "division": division_data,
                "region": region_data,
                "head_office": ho_data
            }
        }, status=status.HTTP_200_OK)

    def _process_simple_offices(self, cases, insurance_company):
        """Process cases for insurance companies with simple office structure (like Tata AIG)"""
        office_data = {}

        for case in cases:
            office_code = case.ins_office_code
            if not office_code:
                continue

            # Get office details
            office = TataAIGOffice.objects.filter(code=office_code, is_active=True).first()
            if not office:
                continue

            # Calculate case amount
            total_case_amount = 0
            for _, prices in case.test_price.items():
                # For Tata AIG, we might use different pricing logic
                # For now, using urban charge as default
                total_case_amount += int(prices.get("lic_urban_charge", 0))

            # Office level
            if office_code not in office_data:
                office_data[office_code] = {"cases": 0, "total_amount": 0, "office_name": office.name}
            office_data[office_code]["cases"] += 1
            office_data[office_code]["total_amount"] += total_case_amount

        return Response({
            "success": True,
            "data": {
                "offices": office_data
            }
        }, status=status.HTTP_200_OK)
